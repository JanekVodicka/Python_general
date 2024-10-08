# EXCEL
# Modul pro excel
import pandas as pd

# Import excelu
excel = pd.read_excel("zk.xlsx", sheet_name = "List1")

# Zobrazení excelu
print(excel.head())	# Prvních 5 řádků
print(excel.tail())	# Posledních 5 řádků

# Vytvoří seznam, dle kterého se dají vybírat sloupce, odpovídá obsahu prvního řádku
Hlavicka = excel.columns
print('První řádek excelu:\n', Hlavicka)
# Přístup k elementu
print('Druhá pozice v prvním řádku:\n', Hlavicka[1])

# Tvorba seznamu - sloupce - dle hlavičky
Jmeno = list(excel[Hlavicka[1]])
print(Jmeno)

# Přístup k prvku
element = excel.iloc[1,1]
print(element)

# Vyškrtne řádky
excel.drop([1,2])

""" List each company with respective product count
 	List products with inventory less than 10
 	List each company with respective total inventory value
 	Write to Spreadsheet: Calculate and write inventory value for each product into spreadsheet"""

# IN GENERAL - implemented functions (Built-in)
# Python has several built-in functions for handling files in general
# io module → create, read, write
# f = open("demofile.txt", "r")
# f.read()
# f.write()
# f.close()
# Python package to work with spreadsheets specifically - for EXCEL
# more practical functions for spreadsheets specifically
# easier to use

# EXCEL Openpyxl
import openpyxl

# Loading/reading excel
inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

# Dictionaries
products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
	supplier_name = product_list.cell(product_row, 4).value
	inventory = product_list.cell(product_row, 2).value
	price = product_list.cell(product_row, 3).value
	product_num = product_list.cell(product_row, 1).value
	inventory_price = product_list.cell(product_row, 5)

	# calculation number of products per supplier - "kolikrát se opakuje název společnosti"
	if supplier_name in products_per_supplier:
		current_num_products = products_per_supplier.get(supplier_name)
		products_per_supplier[supplier_name] = current_num_products + 1
	else:
		products_per_supplier[supplier_name] = 1

	# calculation total value of inventory per supplier
	if supplier_name in total_value_per_supplier:
		current_total_value = total_value_per_supplier.get(supplier_name)
		total_value_per_supplier[supplier_name] = current_total_value + inventory * price
	else:
		total_value_per_supplier[supplier_name] = inventory * price

	# logic products with inventory less then 10
	if inventory < 10:
		products_under_10_inv[int(product_num)] = int(inventory)

	# add value for total inventory price
	inventory_price.value = inventory * price

print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)

# Uložení:
inv_file.save("inventory_with_total_value.xlsx")
